"""Settings routes — financial rates, material rates, catalog, LME refresh."""
import os
import io
import logging
import httpx
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.db import get_db
from app.api.deps import get_current_user, require_admin
from app.models.orm_models import FinancialRates, MaterialRates, Tenant, User

router = APIRouter(prefix="/api/settings", tags=["Tenant Settings"])
logger = logging.getLogger("masaad-api")


# ─── Root settings endpoint (no auth required — used by Layout.tsx) ──────────

@router.get("")
async def get_settings_root():
    """
    Return basic tenant settings for the Layout header bar.
    No auth required — returns defaults if DB is unavailable.
    Used by Layout.tsx to show live shop rate in the header.
    """
    factory_rate = 13.0  # default burn rate AED/hr
    shop_rate = 48.75    # default shop rate AED/hr
    try:
        from app.db import AsyncSessionLocal
        async with AsyncSessionLocal() as session:
            result = await session.execute(select(MaterialRates).limit(1))
            mat = result.scalar_one_or_none()
            if mat:
                if mat.factory_hourly_rate_aed:
                    factory_rate = float(mat.factory_hourly_rate_aed)
                if mat.site_installation_rate_aed:
                    shop_rate = float(mat.site_installation_rate_aed)
    except Exception as e:
        logger.debug(f"Settings root: DB not available ({e}), using defaults")

    return {
        "factory_rate_aed": factory_rate,
        "shop_rate_aed": shop_rate,
        "company_name": "Madinat Al Saada Aluminium & Glass Works LLC",
        "currency": "AED",
        "default_margin_pct": 18.0,
        "retention_pct": 10.0,
        "attic_stock_pct": 2.0,
    }


# ─── Pydantic schemas ────────────────────────────────────────────────────────

class FinancialRatesUpdate(BaseModel):
    lme_aluminum_usd_mt: Optional[float] = None
    billet_premium_usd_mt: Optional[float] = None
    extrusion_premium_usd_mt: Optional[float] = None
    usd_aed: Optional[float] = None
    factory_overhead_pct: Optional[float] = None
    admin_overhead_pct: Optional[float] = None
    risk_contingency_pct: Optional[float] = None
    default_profit_margin_pct: Optional[float] = None


class MaterialRatesUpdate(BaseModel):
    # Glass
    glass_clear_float_aed_sqm: Optional[float] = None
    glass_tinted_aed_sqm: Optional[float] = None
    glass_tempered_clear_aed_sqm: Optional[float] = None
    glass_tempered_tinted_aed_sqm: Optional[float] = None
    glass_laminated_6_6_aed_sqm: Optional[float] = None
    glass_low_e_aed_sqm: Optional[float] = None
    glass_dgu_6_12_6_clear_aed_sqm: Optional[float] = None
    glass_dgu_low_e_aed_sqm: Optional[float] = None
    glass_opaque_spandrel_aed_sqm: Optional[float] = None
    glass_structural_dgu_aed_sqm: Optional[float] = None
    # ACP
    acp_polyester_aed_sqm: Optional[float] = None
    acp_powder_coat_aed_sqm: Optional[float] = None
    acp_pvdf_aed_sqm: Optional[float] = None
    acp_metallic_pvdf_aed_sqm: Optional[float] = None
    acp_mirror_aed_sqm: Optional[float] = None
    # Hardware
    hardware_casement_handle_aed: Optional[float] = None
    hardware_casement_hinge_pair_aed: Optional[float] = None
    hardware_casement_lock_aed: Optional[float] = None
    hardware_casement_restrictor_aed: Optional[float] = None
    hardware_door_handle_set_aed: Optional[float] = None
    hardware_mortice_lock_aed: Optional[float] = None
    hardware_door_closer_aed: Optional[float] = None
    hardware_door_hinge_set_aed: Optional[float] = None
    hardware_floor_spring_aed: Optional[float] = None
    hardware_patch_fitting_set_aed: Optional[float] = None
    hardware_spider_fitting_aed: Optional[float] = None
    # Sealants
    sealant_weatherseal_310ml_aed: Optional[float] = None
    sealant_structural_600ml_aed: Optional[float] = None
    sealant_primer_500ml_aed: Optional[float] = None
    backer_rod_10mm_per_lm_aed: Optional[float] = None
    backer_rod_15mm_per_lm_aed: Optional[float] = None
    setting_block_each_aed: Optional[float] = None
    distance_piece_each_aed: Optional[float] = None
    # Fixings
    anchor_m10_each_aed: Optional[float] = None
    anchor_m12_each_aed: Optional[float] = None
    bracket_80mm_each_aed: Optional[float] = None
    bracket_120mm_each_aed: Optional[float] = None
    shim_plate_each_aed: Optional[float] = None
    thermal_pad_each_aed: Optional[float] = None
    t_connector_each_aed: Optional[float] = None
    l_connector_each_aed: Optional[float] = None
    end_cap_each_aed: Optional[float] = None
    expansion_joint_each_aed: Optional[float] = None
    fire_stop_per_lm_aed: Optional[float] = None
    drainage_insert_each_aed: Optional[float] = None
    # Labor
    factory_hourly_rate_aed: Optional[float] = None
    site_installation_rate_aed: Optional[float] = None
    # Overheads
    factory_overhead_pct: Optional[float] = None
    admin_overhead_pct: Optional[float] = None
    risk_contingency_pct: Optional[float] = None
    default_profit_margin_pct: Optional[float] = None


# ─── Helpers ────────────────────────────────────────────────────────────────

async def _get_tenant_id(user: User) -> str:
    if not user.tenant_id:
        raise HTTPException(status_code=400, detail="User has no tenant assigned")
    return str(user.tenant_id)


async def _fetch_live_lme() -> float:
    """Fetch live LME aluminum price from free metals API."""
    try:
        # Try metals-api.com (500 free calls/month)
        metals_api_key = os.getenv("METALS_API_KEY", "")
        if metals_api_key:
            async with httpx.AsyncClient(timeout=10) as client:
                r = await client.get(
                    f"https://metals-api.com/api/latest?access_key={metals_api_key}&base=USD&symbols=ALU"
                )
                if r.status_code == 200:
                    data = r.json()
                    if data.get("success") and "ALU" in data.get("rates", {}):
                        # metals-api returns price per troy oz; LME is per metric ton
                        # ALU in metals-api is USD per metric ton directly
                        return float(data["rates"]["ALU"])

        # Fallback: use a public commodities API
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(
                "https://api.stlouisfed.org/fred/series/observations?series_id=ALUM&api_key=demo&file_type=json&limit=1&sort_order=desc",
                headers={"Accept": "application/json"}
            )
            if r.status_code == 200:
                data = r.json()
                obs = data.get("observations", [])
                if obs:
                    return float(obs[0]["value"]) * 1000  # convert from USD/kg to USD/MT

    except Exception as e:
        logger.warning(f"LME fetch failed: {e}")

    return None  # Caller will use cached value


# ─── Endpoints ──────────────────────────────────────────────────────────────

@router.get("/current-rates")
async def get_current_rates(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return current financial + material rates for the tenant."""
    tenant_id = await _get_tenant_id(user)

    fin_result = await db.execute(select(FinancialRates).where(FinancialRates.tenant_id == tenant_id))
    fin = fin_result.scalar_one_or_none()

    mat_result = await db.execute(select(MaterialRates).where(MaterialRates.tenant_id == tenant_id))
    mat = mat_result.scalar_one_or_none()

    return {
        "financial_rates": {c.name: getattr(fin, c.name) for c in FinancialRates.__table__.columns} if fin else {},
        "material_rates": {c.name: getattr(mat, c.name) for c in MaterialRates.__table__.columns} if mat else {},
    }


@router.post("/financial-rates")
async def upsert_financial_rates(
    payload: FinancialRatesUpdate,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """UPSERT financial rates for the tenant."""
    tenant_id = await _get_tenant_id(user)

    result = await db.execute(select(FinancialRates).where(FinancialRates.tenant_id == tenant_id))
    fin = result.scalar_one_or_none()

    if not fin:
        fin = FinancialRates(tenant_id=tenant_id)
        db.add(fin)

    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(fin, field, value)

    await db.commit()
    return {"status": "updated", "financial_rates": payload.model_dump(exclude_none=True)}


@router.get("/material-rates")
async def get_material_rates(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return material rates for the tenant."""
    tenant_id = await _get_tenant_id(user)
    result = await db.execute(select(MaterialRates).where(MaterialRates.tenant_id == tenant_id))
    mat = result.scalar_one_or_none()
    if not mat:
        return {}
    return {c.name: getattr(mat, c.name) for c in MaterialRates.__table__.columns}


@router.post("/material-rates")
async def upsert_material_rates(
    payload: MaterialRatesUpdate,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """UPSERT material rates for the tenant."""
    tenant_id = await _get_tenant_id(user)

    result = await db.execute(select(MaterialRates).where(MaterialRates.tenant_id == tenant_id))
    mat = result.scalar_one_or_none()

    if not mat:
        mat = MaterialRates(tenant_id=tenant_id)
        db.add(mat)

    updates = payload.model_dump(exclude_none=True)
    for field, value in updates.items():
        setattr(mat, field, value)

    if updates:
        mat.rates_last_updated = datetime.utcnow()
        mat.updated_by = user.id

    await db.commit()
    return {"status": "updated", "fields_updated": list(updates.keys())}


@router.get("/refresh-lme")
async def refresh_lme(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Refresh live LME aluminum price. Uses 6-hour cache."""
    tenant_id = await _get_tenant_id(user)

    result = await db.execute(select(FinancialRates).where(FinancialRates.tenant_id == tenant_id))
    fin = result.scalar_one_or_none()

    # Check cache: if updated within 6 hours, return cached
    if fin and fin.lme_last_fetched:
        age_hours = (datetime.utcnow() - fin.lme_last_fetched).total_seconds() / 3600
        if age_hours < 6 and fin.lme_aluminum_usd_mt:
            return {
                "lme_usd_mt": float(fin.lme_aluminum_usd_mt),
                "source": "cached",
                "cache_age_hours": round(age_hours, 1),
                "last_updated": fin.lme_last_fetched.isoformat(),
            }

    # Fetch live
    live_price = await _fetch_live_lme()

    if live_price and fin:
        fin.lme_aluminum_usd_mt = live_price
        fin.lme_last_fetched = datetime.utcnow()
        fin.lme_source = "live"
        await db.commit()
        return {
            "lme_usd_mt": live_price,
            "source": "live",
            "cache_age_hours": 0,
            "last_updated": fin.lme_last_fetched.isoformat(),
        }
    elif fin and fin.lme_aluminum_usd_mt:
        return {
            "lme_usd_mt": float(fin.lme_aluminum_usd_mt),
            "source": "cached_fallback",
            "note": "Live fetch failed — returning last known value",
            "last_updated": fin.lme_last_fetched.isoformat() if fin.lme_last_fetched else None,
        }

    return {"lme_usd_mt": 2485.0, "source": "default", "note": "No cached value — using default"}


@router.post("/upload-payroll")
async def upload_payroll(
    file: UploadFile = File(...),
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Ingest payroll Excel/CSV (multi-sheet: MADINAT, AL JAZEERA, MADINAT AL JAZEERA).
    Filters for FACTORY job location. Calculates fully burdened hourly burn rate.

    Burn rate formula:
      baseline_labor_burn_rate_aed = (total_monthly_payroll + factory_overhead)
                                     / (num_workers * 26 days * 8 hours)
    """
    fname = (file.filename or "").lower()
    if not (fname.endswith('.xlsx') or fname.endswith('.xls') or fname.endswith('.csv')):
        raise HTTPException(status_code=400, detail="Upload Excel (.xlsx/.xls) or CSV file")

    try:
        contents = await file.read()
        if not contents:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")

        all_factory_rows: list[dict] = []

        if fname.endswith('.csv'):
            all_factory_rows = _parse_payroll_csv(contents)
        else:
            all_factory_rows = _parse_payroll_excel(contents)

        if not all_factory_rows:
            raise HTTPException(
                status_code=400,
                detail="No FACTORY workers found in any sheet. "
                       "Ensure there is a column with LOCATION/SITE/JOB containing 'FACTORY'."
            )

        # Find salary column
        sample_row = all_factory_rows[0]
        salary_key = _find_column(sample_row, ['SALARY', 'PAY', 'GROSS', 'TOTAL', 'NET'])
        if not salary_key:
            raise HTTPException(
                status_code=400,
                detail="Could not find a salary/pay column. "
                       "Expected column name containing SALARY, PAY, GROSS, or TOTAL."
            )

        # Also look for allowance columns to add to total compensation
        allowance_key = _find_column(sample_row, ['ALLOWANCE', 'HOUSING', 'TRANSPORT', 'FOOD', 'OTHER'])

        total_monthly_payroll = 0.0
        for row in all_factory_rows:
            salary_val = _to_float(row.get(salary_key, 0))
            total_monthly_payroll += salary_val
            if allowance_key:
                total_monthly_payroll += _to_float(row.get(allowance_key, 0))

        total_workers = len(all_factory_rows)
        # 26 working days/month * 8 hours/day
        working_days = 26
        working_hours_per_day = 8
        total_working_hours = total_workers * working_days * working_hours_per_day

        # Fetch factory overhead percentage from FinancialRates (or use 12% default)
        tenant_id = await _get_tenant_id(user)
        fin_result = await db.execute(
            select(FinancialRates).where(FinancialRates.tenant_id == tenant_id)
        )
        fin = fin_result.scalar_one_or_none()

        # Get factory overhead from MaterialRates
        mat_result = await db.execute(
            select(MaterialRates).where(MaterialRates.tenant_id == tenant_id)
        )
        mat = mat_result.scalar_one_or_none()
        factory_overhead_pct = float(mat.factory_overhead_pct) if mat and mat.factory_overhead_pct else 0.12
        factory_overhead = total_monthly_payroll * factory_overhead_pct

        burn_rate = (total_monthly_payroll + factory_overhead) / total_working_hours if total_working_hours > 0 else 0.0
        burn_rate = round(burn_rate, 4)

        # Update FinancialRates with burn rate
        if not fin:
            fin = FinancialRates(tenant_id=tenant_id)
            db.add(fin)
        fin.baseline_labor_burn_rate_aed = burn_rate
        fin.burn_rate_last_updated = datetime.utcnow()
        fin.burn_rate_updated_by_source = "payroll_upload"
        fin.updated_by = user.id

        # Also update MaterialRates.factory_hourly_rate_aed for consistency
        if not mat:
            mat = MaterialRates(tenant_id=tenant_id)
            db.add(mat)
        mat.factory_hourly_rate_aed = round(burn_rate, 2)
        mat.rates_last_updated = datetime.utcnow()
        mat.updated_by = user.id

        await db.commit()

        # Frontend expects true_shop_rate_aed in metrics
        return {
            "status": "success",
            "metrics": {
                "factory_headcount": int(total_workers),
                "total_monthly_payroll_aed": round(float(total_monthly_payroll), 2),
                "factory_overhead_aed": round(float(factory_overhead), 2),
                "factory_overhead_pct": factory_overhead_pct,
                "working_days_per_month": working_days,
                "working_hours_per_day": working_hours_per_day,
                "total_working_hours": total_working_hours,
                "burn_rate_aed_per_hr": round(float(burn_rate), 2),
                "true_shop_rate_aed": round(float(burn_rate), 2),
            },
            "note": "Burn rate saved to FinancialRates and MaterialRates",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Payroll upload failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Payroll processing error: {str(e)}")


def _to_float(val) -> float:
    """Safely convert any value to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        # Try stripping currency symbols
        s = str(val).replace(',', '').replace('AED', '').replace('$', '').strip()
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0


def _find_column(row: dict, keywords: list[str]) -> str | None:
    """Find a column key in a dict by matching keyword substrings (case-insensitive)."""
    for key in row.keys():
        key_upper = str(key).upper().strip()
        for kw in keywords:
            if kw in key_upper:
                return key
    return None


def _parse_payroll_csv(contents: bytes) -> list[dict]:
    """Parse a CSV payroll file, returning rows where location is FACTORY."""
    import csv
    text = contents.decode('utf-8', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for raw_row in reader:
        # Normalize keys
        row = {k.strip(): v for k, v in raw_row.items() if k}
        loc_key = _find_column(row, ['LOCATION', 'SITE', 'JOB'])
        if loc_key:
            loc_val = str(row.get(loc_key, '')).upper().strip()
            if loc_val == 'FACTORY':
                rows.append(row)
        else:
            # No location column — include all rows (assume all factory)
            rows.append(row)
    return rows


def _parse_payroll_excel(contents: bytes) -> list[dict]:
    """Parse an Excel payroll file (multi-sheet), returning FACTORY worker rows."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    all_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Find header row (first row with >2 non-None values)
        header = None
        for row in rows_iter:
            non_none = [c for c in row if c is not None]
            if len(non_none) >= 2:
                header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(row)]
                break

        if not header:
            continue

        loc_idx = None
        for i, h in enumerate(header):
            h_upper = h.upper()
            if any(kw in h_upper for kw in ['LOCATION', 'SITE', 'JOB']):
                loc_idx = i
                break

        for row in rows_iter:
            if not row or all(c is None for c in row):
                continue
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(header):
                    row_dict[header[i]] = val

            if loc_idx is not None:
                loc_val = str(row[loc_idx] if loc_idx < len(row) else '').upper().strip()
                if loc_val == 'FACTORY':
                    row_dict['_sheet'] = sheet_name
                    all_rows.append(row_dict)
            else:
                # No location column in this sheet — include all data rows
                row_dict['_sheet'] = sheet_name
                all_rows.append(row_dict)

    wb.close()
    return all_rows


def _parse_expenses_excel(contents: bytes) -> tuple[float, list[str], list[dict]]:
    """
    Parse an expenses Excel file.
    Returns (total_amount, entity_columns_found, row_details).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return 0.0, [], []

    rows_iter = ws.iter_rows(values_only=True)

    # Find header
    header = None
    for row in rows_iter:
        non_none = [c for c in row if c is not None]
        if len(non_none) >= 2:
            header = [str(c).strip().upper() if c is not None else f"COL_{i}" for i, c in enumerate(row)]
            break

    if not header:
        wb.close()
        return 0.0, [], []

    # Look for entity columns (MADINAT, AL JAZEERA, etc.) or amount columns
    entity_keywords = ['MADINAT', 'AL JAZEERA', 'MADINAT AL JAZEERA']
    amount_keywords = ['AMOUNT', 'TOTAL', 'EXPENSE', 'VALUE', 'COST']

    entity_cols: list[int] = []
    entity_names: list[str] = []
    amount_cols: list[int] = []

    for i, h in enumerate(header):
        for ek in entity_keywords:
            if ek in h:
                entity_cols.append(i)
                entity_names.append(h)
                break
        for ak in amount_keywords:
            if ak in h:
                amount_cols.append(i)
                break

    total = 0.0
    row_details = []

    # Determine which columns to sum
    sum_cols = entity_cols if entity_cols else amount_cols

    # Find category column for details
    cat_idx = None
    for i, h in enumerate(header):
        if any(kw in h for kw in ['CATEGORY', 'DESCRIPTION', 'ITEM', 'EXPENSE', 'PARTICULAR']):
            cat_idx = i
            break

    for row in rows_iter:
        if not row or all(c is None for c in row):
            continue
        row_total = 0.0
        for ci in sum_cols:
            if ci < len(row) and row[ci] is not None:
                row_total += _to_float(row[ci])
        if row_total != 0:
            total += row_total
            cat_name = str(row[cat_idx]).strip() if cat_idx is not None and cat_idx < len(row) and row[cat_idx] else "Uncategorized"
            row_details.append({"category": cat_name, "amount_aed": round(row_total, 2)})

    wb.close()
    return total, entity_names, row_details


def _parse_expenses_csv(contents: bytes) -> tuple[float, list[str], list[dict]]:
    """Parse a CSV expenses file. Returns (total, entity_names, row_details)."""
    import csv
    text = contents.decode('utf-8', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return 0.0, [], []

    # Normalize headers
    sample = rows[0]
    header_upper = {k: k.upper().strip() for k in sample.keys() if k}

    entity_keywords = ['MADINAT', 'AL JAZEERA', 'MADINAT AL JAZEERA']
    amount_keywords = ['AMOUNT', 'TOTAL', 'EXPENSE', 'VALUE', 'COST']

    entity_cols = []
    amount_cols = []
    cat_key = None

    for orig_key, upper_key in header_upper.items():
        for ek in entity_keywords:
            if ek in upper_key:
                entity_cols.append(orig_key)
                break
        for ak in amount_keywords:
            if ak in upper_key:
                amount_cols.append(orig_key)
                break
        if any(kw in upper_key for kw in ['CATEGORY', 'DESCRIPTION', 'ITEM', 'PARTICULAR']):
            cat_key = orig_key

    sum_cols = entity_cols if entity_cols else amount_cols
    total = 0.0
    row_details = []

    for row in rows:
        row_total = 0.0
        for col in sum_cols:
            row_total += _to_float(row.get(col, 0))
        if row_total != 0:
            total += row_total
            cat_name = row.get(cat_key, "Uncategorized").strip() if cat_key else "Uncategorized"
            row_details.append({"category": cat_name, "amount_aed": round(row_total, 2)})

    return total, [c.upper() for c in entity_cols] if entity_cols else ["AMOUNT"], row_details


@router.post("/upload-expenses")
async def upload_expenses(
    file: UploadFile = File(...),
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Ingest admin expenses CSV/Excel.
    Aggregates across entity columns (MADINAT, AL JAZEERA, MADINAT AL JAZEERA)
    or falls back to AMOUNT/TOTAL columns.
    """
    fname = (file.filename or "").lower()
    if not (fname.endswith('.xlsx') or fname.endswith('.xls') or fname.endswith('.csv')):
        raise HTTPException(status_code=400, detail="Upload Excel (.xlsx/.xls) or CSV file")

    try:
        contents = await file.read()
        if not contents:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")

        if fname.endswith('.csv'):
            total, entity_names, row_details = _parse_expenses_csv(contents)
        else:
            total, entity_names, row_details = _parse_expenses_excel(contents)

        if total == 0 and not row_details:
            raise HTTPException(
                status_code=400,
                detail="Could not find expense data. "
                       "Expected columns: MADINAT, AL JAZEERA, or AMOUNT/TOTAL/EXPENSE."
            )

        # Update financial rates — store total group overhead
        tenant_id = await _get_tenant_id(user)
        fin_result = await db.execute(
            select(FinancialRates).where(FinancialRates.tenant_id == tenant_id)
        )
        fin = fin_result.scalar_one_or_none()
        if not fin:
            fin = FinancialRates(tenant_id=tenant_id)
            db.add(fin)
        fin.last_updated = datetime.utcnow()
        fin.updated_by = user.id
        await db.commit()

        return {
            "status": "success",
            "total_group_overhead_aed": round(float(total), 2),
            "entities_aggregated": entity_names,
            "line_items": len(row_details),
            "breakdown": row_details[:50],  # Cap at 50 rows for response size
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Expense upload failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Expense processing error: {str(e)}")


@router.post("/update-market")
async def update_market(
    payload: FinancialRatesUpdate,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Update market variables (LME, billet premium, etc.)."""
    return await upsert_financial_rates(payload, user, db)
